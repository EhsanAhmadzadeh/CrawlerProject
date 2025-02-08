import os
import logging
import pandas as pd
from openpyxl import load_workbook
from config import AppConfig


def create_excel_if_not_exists():
    """Ensures the Excel file is initialized with correct headers."""
    if not os.path.exists(AppConfig.EXCEL_FILE):
        with pd.ExcelWriter(AppConfig.EXCEL_FILE, engine="openpyxl") as writer:
            pd.DataFrame(
                columns=[
                    "app_id", "app_name", "description_content",
                    "installation_counts", "app_score", "app_category",
                    "app_size", "app_last_update", "app_images"
                ]
            ).to_excel(writer, sheet_name="Apps", index=False)

            pd.DataFrame(
                columns=[
                    "comment_id", "app_id", "username", "account_id",
                    "rating", "comment", "comment_date"
                ]
            ).to_excel(writer, sheet_name="Comments", index=False)

        logging.info(f"📁 Created new Excel file: {AppConfig.EXCEL_FILE}")


def write_to_excel(app_df: pd.DataFrame, comments_df: pd.DataFrame):
    """Writes app data and comments to an Excel file, appending rows without duplicating headers."""
    file_exists = os.path.exists(AppConfig.EXCEL_FILE)

    with pd.ExcelWriter(AppConfig.EXCEL_FILE, mode="a", engine="openpyxl", if_sheet_exists="overlay") as writer:
        if file_exists:
            book = load_workbook(AppConfig.EXCEL_FILE)

            # If these sheets exist, find the next blank row
            if "Apps" in book.sheetnames:
                app_sheet = book["Apps"]
                app_start_row = app_sheet.max_row + 0  # next row
            else:
                app_start_row = 0

            if "Comments" in book.sheetnames:
                comment_sheet = book["Comments"]
                comment_start_row = comment_sheet.max_row + 0
            else:
                comment_start_row = 0

            # Write the data without headers, starting at next row
            app_df.to_excel(
                writer, sheet_name="Apps", index=False,
                header=False, startrow=app_start_row
            )
            comments_df.to_excel(
                writer, sheet_name="Comments", index=False,
                header=False, startrow=comment_start_row
            )
        else:
            # If the file doesn't exist for some reason, write with headers
            app_df.to_excel(writer, sheet_name="Apps", index=False)
            comments_df.to_excel(writer, sheet_name="Comments", index=False)
