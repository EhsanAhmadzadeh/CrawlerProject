from config import AppConfig
from openpyxl import load_workbook
import pandas as pd
import logging

def create_excel_if_not_exists():
    """Ensures the Excel file is initialized with correct headers."""
    import os
    if not os.path.exists(AppConfig.EXCEL_FILE):
        with pd.ExcelWriter(AppConfig.EXCEL_FILE, engine="openpyxl") as writer:
            pd.DataFrame(columns=["app_id", "app_name", "description_content", "installation_counts", "app_score", "app_category", "app_size", "app_last_update", "app_images"]).to_excel(writer, sheet_name="Apps", index=False)
            pd.DataFrame(columns=["comment_id", "app_id", "username", "account_id", "rating", "comment", "comment_date"]).to_excel(writer, sheet_name="Comments", index=False)
        logging.info(f"📁 Created new Excel file: {AppConfig.EXCEL_FILE}")
        
        
def write_to_excel(app_df, comments_df):
    """Writes data to an Excel file without duplicating apps."""
    import os
    file_exists = os.path.exists(AppConfig.EXCEL_FILE)

    with pd.ExcelWriter(AppConfig.EXCEL_FILE, mode="a", engine="openpyxl", if_sheet_exists="overlay") as writer:
        if file_exists:
            book = load_workbook(AppConfig.EXCEL_FILE)
            app_start_row = book["Apps"].max_row if "Apps" in book.sheetnames else 1
            comment_start_row = book["Comments"].max_row if "Comments" in book.sheetnames else 1

            app_df.to_excel(writer, sheet_name="Apps", index=False, header=False, startrow=app_start_row)
            comments_df.to_excel(writer, sheet_name="Comments", index=False, header=False, startrow=comment_start_row)

        else:
            app_df.to_excel(writer, sheet_name="Apps", index=False)
            comments_df.to_excel(writer, sheet_name="Comments", index=False)
