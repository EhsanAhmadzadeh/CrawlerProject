import logging
import os
import asyncio
import pandas as pd
from openpyxl import load_workbook
from services import get_app_links, get_app_metadata, get_all_comments_page_html, get_comments_data

# Configure Logging
LOG_FILE = "output/crawler.log"
os.makedirs("output", exist_ok=True)  # Ensure output folder exists

logging.basicConfig(
    filename=LOG_FILE,  # Save logs to file
    level=logging.INFO,  # Log level
    format="%(asctime)s - %(levelname)s - %(message)s",  # Log format
    filemode="a"
)

console_handler = logging.StreamHandler()  # Also show logs in console
console_handler.setLevel(logging.INFO)
console_handler.setFormatter(logging.Formatter("%(asctime)s - %(levelname)s - %(message)s"))
logging.getLogger().addHandler(console_handler)

# Excel File Path
EXCEL_FILE = "output/apps_data.xlsx"

def create_excel_if_not_exists():
    """Ensures the Excel file is initialized with correct headers."""
    if not os.path.exists(EXCEL_FILE):
        with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl") as writer:
            pd.DataFrame(columns=["app_id", "app_name", "description_content", "installation_counts", "app_score", "app_category", "app_size", "app_last_update", "app_images"]).to_excel(writer, sheet_name="Apps", index=False)
            pd.DataFrame(columns=["comment_id", "app_id", "username", "account_id", "rating", "comment", "comment_date"]).to_excel(writer, sheet_name="Comments", index=False)
        logging.info(f"📁 Created new Excel file: {EXCEL_FILE}")

async def process_app(full_url):
    """Processes each app and writes data to Excel."""

    logging.info(f"🔍 Fetching metadata for: {full_url}")
    app_metadata = get_app_metadata(full_url)
    if not app_metadata:
        logging.warning(f"❌ Skipping {full_url} (No metadata found)")
        return
    
    logging.info(f"✅ App metadata fetched: {app_metadata['app_name']} (ID: {app_metadata['app_id']})")

    logging.info(f"🔄 Scraping comments for {app_metadata['app_name']}...")
    page_html = await get_all_comments_page_html(full_url)
    comments = get_comments_data(page_html, app_metadata["app_id"])
    
    logging.info(f"💬 Fetched {len(comments)} comments for {app_metadata['app_name']}")

    app_df = pd.DataFrame([app_metadata])
    comments_df = pd.DataFrame(comments)

    write_to_excel(app_df, comments_df)

    logging.info(f"📂 Data saved successfully for: {app_metadata['app_name']}")
    logging.info("------------------------------------------------------------")

def write_to_excel(app_df, comments_df):
    """Writes data to an Excel file without duplicating apps."""

    file_exists = os.path.exists(EXCEL_FILE)

    with pd.ExcelWriter(EXCEL_FILE, mode="a", engine="openpyxl", if_sheet_exists="overlay") as writer:
        if file_exists:
            book = load_workbook(EXCEL_FILE)
            
            # Find the correct row positions
            app_start_row = book["Apps"].max_row if "Apps" in book.sheetnames else 1
            comment_start_row = book["Comments"].max_row if "Comments" in book.sheetnames else 1

            app_df.to_excel(writer, sheet_name="Apps", index=False, header=False, startrow=app_start_row)
            comments_df.to_excel(writer, sheet_name="Comments", index=False, header=False, startrow=comment_start_row)

        else:
            app_df.to_excel(writer, sheet_name="Apps", index=False)
            comments_df.to_excel(writer, sheet_name="Comments", index=False)

async def main():
    logging.info("🚀 Starting app data collection...")
    create_excel_if_not_exists()

    MAIN_DOMAIN = "https://cafebazaar.ir"
    MENTAL_HEALTH_APP_ROUTE = "/lists/ml-mental-health-exercises"
    
    links = get_app_links(MAIN_DOMAIN + MENTAL_HEALTH_APP_ROUTE)
    logging.info(f"🔗 Found {len(links)} apps to process")

    for link in links:
        await process_app(MAIN_DOMAIN + link)

    logging.info("✅ All apps processed successfully!")

asyncio.run(main())
